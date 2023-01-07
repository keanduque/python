"""
Practice Code and Notes
"""

# print('Hello World Python')
# print('*' * 10)

# price = 10
# rating = 4.9
# name = 'Mosh'
# is_published = False
#
# # ex1
# patient_name = "John Smith"
# age = 20
# is_new_patient = False

# ex2
# name = input('What is your name? ')
# fav_color = input('What is your favorite color? ')
# print(name + " likes "+ fav_color)

# Convert
# birth_year = input('Birth year: ')
# print(type(birth_year))
# age = 2022 - int(birth_year)
# print(type(age))
# print(age)

# ex3
# weight_lbs = input("Enter your weight (in lbs) : ")
# weight_kg = float(weight_lbs) * 0.453
# print("your weight in (KG) is : " + str(weight_kg))

# -------------------------------------------------------------

text = ''' 
Hi John,

Here is our first email to you.

Thank you,
the Support team 
'''
# print(text)

text2 = 'Python for Beginners'
#  0123456789 AndSoOn-1
# print(text2[-2]) # r
# print(text2[0:3]) # Pyt
# print(text2[1:]) # ython for Beginners
# print(text2[:5]) # Pytho
# print(text2[:]) # return all the characters
# name = 'Jennifer'
# print(name[1:-1]) # ennife

first = 'Kean'
last = 'Duque'
message = first + ' [' + last + '] is a coder'  # ideal formatting

# f is for formatted strings
msg = f'{first} [{last}] is a coder'  # formatted strings for python defining placeholder
# print(msg)

course = 'Python for Beginners'
# print(len(course)) #20 start with 1
# print(course)
# print(course.upper())
# print(course.lower())
# print(course.title())
# *****find method is return the index or sequence of char
# print(course.find('Beginners')) #index is 0 case sensitive - result 11
# print(course.find('o')) #4
# print(course.find('O')) #-1
#
# *****replacing words from string
# print(course.replace('Beginners', 'Absolute Beginners'))

# *****produces boolean
# print('Python' in course) # True
# print('python' in course) # False

# print(10 // 3) # 3
# print(10 / 3) # 3.3333333333333335
# print(10 ** 3) # 1000
# print(10 % 3) # 1

# Augmented assignment operator +=
# x = 10
# x = x + 3 # 13
# x += 3 # 16 incrementing a number augmented ass ops
# print(x)

# Operator Presedence
# x = (10 + 3) * 2 ** 2 # = 34 PEMDAS rules

# Math Functions
# x = 2.9
# print(round(x)) #3
# print(abs(-2.9)) # 2.9 absolute values

# you can import math module in python for complex computations

# print(math.ceil((2.9))) #3
# print(math.floor(2.9)) #2

# -------------------------------------------------------------

# IF ELSE CONDITIONS
is_hot = False
is_cold = True

# if is_hot:
#     print("It's a hot day")
#     print("Drink plenty of water")
# elif is_cold:
#     print("It's a cold day")
#     print("Wear warm clothes")
# else:
#     print("It's a lovely day")
# print("Enjoy your day")

price_house = 1000000
buyer_good_credit = False

if buyer_good_credit:
    downpayment = price_house * 0.10
else:
    downpayment = price_house * 0.20
# print(f"DP is : ${downpayment}")
# print("Total Price : ${:,}".format(price_house - downpayment))

has_high_income = False
has_good_credit = True
has_criminal_record = False
# if has_good_credit or has_high_income:
#     print("Eligible for loan")
# if has_good_credit and not has_criminal_record:
#     print("Eligible for loan")

temperature = 30
# if temperature > 30:
#     print("It's a hot day")
# else:
#     print("It's not a hot day")

# exercise
# name_input = input("Input your Name:")
# if len(name_input) < 3:
#     print("name must be at least 3 chars")
# elif len(name_input) > 50:
#     print("name can be a max of 50 chars")
# else:
#     print(f'name looks good {name_input}')

# -------------------------------------------------------------

# While loop

guess_count = 1
# while i <= 5:
#     print('*' * i)
#     i = i + 1
# print("Done")

# for item in "Python":
#     print(item)

# for item in ["kean", "cherry", "johnSmith"]:
#     print(item)
#
# for item in [1,2,3,4]:
#     print(item)

# for item in range(10): 0-9
#     print(item)

# for item in range(5, 10): # 5,6,7,8,9
#     print(item)

# for item in range(5, 10, 2): # 5,7,9
#     print(item)

# prices = [10,20,30,1,2,5]
# total_prices = 0
# for price in prices:
#     total_prices += price
#     print(price)
# print(f'total price is :{total_prices}')

# Nested Loops

# for x in range(4):
#     for y in range(3):
#         print(f'({x},{y})')

# Letter F and L nested Loop
# numbers = [5, 2, 5, 2, 2]
# for row in numbers:
#     output = ''
#     for col in range(row):
#         output += 'x'
#     print(output)
# print("\n")
# numberL = [2,2,2,2,5]
# for x in numberL:
#     output = ''
#     for y in range(x):
#         output += 'y'
#     print(output)

# -------------------------------------------------------------

# *****List

# names = ['Kean', 'Cherry', 'Alpha', 'Beta', 'Charlie', 'John', 'Pedro', 'Juan']

# Find the largest Number
# numbers = [2,4,5,7,1,2,10,23,27,8,3]

# large_number = numbers[0]
# for x in numbers:
#     if x > large_number:
#         large_number = x
# print(large_number)

# 2D List

matrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9],
]

# for row in matrix:
#     for item in row:
#       print(item)

# List methods

# numbers = [5, 2, 1, 5, 7, 4]
# print(numbers.index(5)) # to check the position occurence in the list
# numbers.append(27)      # [5, 2, 1, 7, 4, 27]
# numbers.insert(0,10)    # [10, 5, 2, 1, 7, 4, 27]
# numbers.remove(4)       # [10, 5, 2, 1, 7, 27]
# # numbers.clear()       # remove all the items []
# numbers.pop()           # [10, 5, 2, 1, 7]
# print(numbers)
# print(50 in numbers)    # False
# print(5 in numbers)     # True
# print(numbers.count(5)) # count occurences on the lists return repeat 2
# numbers.sort()
# print(numbers)          # [1, 2, 5, 5, 7, 10] ASC
# numbers.reverse()
# print(numbers)          # [10, 7, 5, 5, 2, 1] DESC
# print("")
# numbers2 = numbers.copy()  # [10, 7, 5, 5, 2, 1]
# numbers2.append(10)
# print(numbers2)            # [10, 7, 5, 5, 2, 1, 10]

# ** Exercise List Method
# Remove duplicate in the List
# numbers = [2,2,4,6,3,4,6,1,8,2,4,5]
# uniques = []
# for number in numbers:
#     if number not in uniques:
#         uniques.append(number)
# print(uniques)

# *** Fibonacci Sequence
# num1 = 0
# num2 = 1
# for x in range(20) :
#     print(num1)
#     fn = num1 + num2
#     num1 = num2
#     num2 = fn

# -------------------------------------------------------------
# ***Tuples - same as list but its immutable
# numbers = (1,2,3)
# print(numbers[0])


# ***Unpacking
coordinates = (1, 2, 3)

# x = coordinates[0]
# y = coordinates[1]
# z = coordinates[2]

# x, y, z = coordinates   # same at the x y z line shorthand
# print(y) # 2

# ***Dictionaries key value pairs
# customer = {
#     "name": "John Smith",
#     "age": 30,
#     "is_verified": True
# }
# print(customer["name"]) # John Smith
# print(customer.get("name")) # John Smith
# print(customer.get("birthdate")) # None

# # update values using get
# customer["name"] = "Jack Smith"
# print(customer["name"]) # Jack Smith
# print(customer.get("birthdate", "Jan 1 1980")) # Jan 1 1980
# print("")

# # add new key value pairs in the dictionary or object
# customer["hobbies"] = "Basketball"
# customer["position"] = "Software Engineer"
# print(customer)

# Excercise for dictionary
# Translate Number to Word
# phone = input("Phone: ")
# digit_map = {
#     "+" : "(Plus)",
#     "1" : "One",
#     "2" : "Two",
#     "3" : "Three",
#     "4" : "Four",
#     "5" : "Five",
#     "6" : "Six",
#     "7" : "Seven",
#     "8" : "Eight",
#     "9" : "Nine",
#     "0" : "Zero"
# }
# output = ""
# for number in phone:
#     output += digit_map.get(number) + " "
# print(output)

# Emoji Converter
# message = input(">")
# words = message.split(" ")
# emojis = {
#     ":)" : "😀",
#     ":(" : "😞"
# }
# out = ""
#
# for word in words:
#     out += emojis.get(word, word) + " "
# print(out)
#


# -------------------------------------------------------------
# **** FUNCTIONS and Parameters
# define a function
def greet_user(first_name, last_name):
    print(f"hi there {first_name} {last_name}")
# greet_user("John","Smith")
# greet_user("Kean", "Duque")

# Keyword Arguments for readability of the code
# this can be use after the positional arguments not before the keyword
# ex: greet_user(first_name="John", "Smith") *** this will throw error
#     greet_user("John",last_name="Smith")       correct usage
# def greet_user(first_name, last_name):
#     print(f"hi there {first_name} {last_name}")
# greet_user(first_name="John",last_name="Smith") # positional arguments
# greet_user(last_name="SmithX", first_name="Johnkeyword") # keyword arguments
# calc_cost(total=50, shipping=5, discount=0.1) #keyword args


# all funtion in python return None
# Return statement
def square(number):
    return number * number

# print(square(2))

# Exceptions try except
# try:
#     age = int(input("Age: "))
#     income = 20000
#     risk = income / age
#     print(age)
# except ZeroDivisionError:
#     print("Age cannot be 0")
# except ValueError:
#     print("Invalid value")
#


# -------------------------------------------------------------
# Classes
class Point:
    def __init__(self, x, y):
        self.x = x
        self.y = y

    @staticmethod
    def move():
        print("move")

    @staticmethod
    def draw():
        print("draw")

# *****Instance
# point1 = Point()
# point1.move()
# ## attributes
# point1.x = 10
# point1.y = 20
# print(point1.x)
# point1.draw()
# print("")
# point2 = Point()
# point2.move()

# ## attributes
# point2.x = 1
# print(point2.x)


# *****Constructor
point = Point(10, 20)
# print(point.x)
# print(point.y)


# Exercise for Class Constructor

class Person:
    def __init__(self, name):
        self.name = name

    def talk(self):
        print(f"hello im {self.name}")


person1 = Person("Kean Duque")
# person1.talk()

person2 = Person("Cherry Quilates")
# person2.talk()


# Inheritance
class Mammal:
    def __init__(self, name):
        self.name = name

    def walk(self):
        print(f"walking {self.name}")


# python dont want empty class
class Dog(Mammal):
    @staticmethod
    def bark():
        print("bark")
    # pass  # for empty class use if classes dont have method


class Cat(Mammal):
    @staticmethod
    def be_annoying():
        print("annoying")


class Rabbit(Mammal):
    pass


dog = Dog("Dog")
# dog.walk()
# dog.bark()

cat = Cat("Cat")
# cat.walk()
# cat.be_annoying()

rabbit = Rabbit("Rabbit")
# rabbit.walk()

# -------------------------------------------------------------

# Modules transfer to converters.py
# to organize codes break code into multiple files
# import converters
from converters import kg_to_lbs  # specific import modules


kg_to_lbs(100)

# print(converters.kg_to_lbs(70))

# Exercise
# import utils
from utils import find_largest_number

numbers = [10, 3, 6, 2, 27]
max_num = find_largest_number(numbers)
# print(max_num)
# print(find_largest_number([10,3,6,2,27]))


# ##Packages
# import ecommerce.shipping
# from ecommerce.shipping import calc_shipping
# from ecommerce import shipping
# shipping.calc_shipping()


# -------------------------------------------------------------

# Generating Random Values

import random

# for i in range(3):
# print(random.random())
# print(random.randint(10, 20))

members = ['Kean', 'Cherry', 'Conan', 'Gudetama']
leader = random.choice(members)
# print(leader)

# -------------------------------------------------------------


# Working with Directories
from pathlib import Path

# Absolute path   - start root of the hard disk.
# ex c:\Program Files\Microsoft
# /usr/local/bin
# Relative path   - current dir to go somewhere else

# path = Path("ecommerce")
path = Path()
# print(path.exists())  # True
# path = Path("email")
# path = Path("email")
# print(path.mkdir())
# print(path.rmdir())

for file in path.glob('*.py'):  # search file using pattern <generator object Path.glob at 0x10aad3ef0>
    print(file)

# -------------------------------------------------------------

"""
Pypi and Pip - Python Package Index
# https://pypi.org/

# pip install openpyxl - done
"""

# -------------------------------------------------------------

# Project1 Automation Excel File
import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']

cell = sheet['a1']
cell = sheet.cell(1, 1)
# print(cell)  # <Cell 'Sheet1'.A1>
# print(cell.value)  # transaction_id
# print(sheet.max_row)  # 4


# -------------------------------------------------------------
"""
Machine Learning is subset of AI
- technique of complex problems example scanning dog cats images
- build model or engines and input lots of data based on the problems you want to solve
examples : Self-driving cars, robotics, language processing, vision processing, forecasting,
stock market trends and weather and games and so on.

Steps :
1. Import the Data
2. Clean the Data
3. Split the data into Training/Test Sets
4. Create a Model ex. Decision trees, Linear, Random Forest etc.
5. Train the Model
6. Make Predictions
7. Evaluate and Improve

Libraries & Tools
- Numpy        - provides multidimensional array
- Pandas       - data analysis lib provide concept called data frame. 2 dimensional similar to an excel spreadsheets
- MatPlotLib   - 2 dimensional plotting libs for creating graphs and paths
- Scikit-Learn - most popular provides all common algo like decision tress, neural networks

Tools : 
https://jupyter.org/ for Machine Learning Project - environment for writing code
- inspecting data
platform : 
install : https://www.anaconda.com/
dataset : https://kaggle.com   

go to terminal and enter : jupyter notebook
it will open : http://localhost:8888/tree on browser
if theres error : use jupyter nbclassic

"""






