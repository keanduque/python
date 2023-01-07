"""Module documentation goes here
# Created By  : Kean Duque
# Created Date: 2022/12/26
# Description : main app
"""
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

import utils
from utils import Dice


# weight = int(input("Enter Weight:"))
# utils.lk_converter(weight)

# utils.guessing_number()

# utils.engine_help()

# utils.find_largest_number(numbers=[2,127,80,6,4,5,7,1,2,10,23,27,8,3])

# utils.remove_duplicates(numbers=[2,2,4,6,3,4,6,1,8,2,4,5])

# utils.fibonacci(20)

# phone = input("Phone: ")
# utils.number_to_word(phone)

# message = input(">")
# print(utils.emoji_converter(message))

# Dice Roll tuple of two random values (0, 0)
# dice = Dice()
# print(dice.roll())


# Project_1 Automation Excel File
# update 4th column for corrected_price
def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    cell4_header = sheet.cell(1, 4)
    cell4_header.value = "corrected_price"

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9

        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,                # start at Row 2
                       max_row=sheet.max_row,    # max row is 4
                       min_col=4,                # column D2
                       max_col=4)                # column D4

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)

# process_workbook('transactions.xlsx')


# FizzBuzz
# divisible by 3 = fizz - divided by 3 or in 3x multiplication table no remainder
# divisible by 5 = buzz - ends in a 5 or 0  or no remainder
# divisible by 3 and 5 = fizz_buzz
#
def fizz_buzz(num):
    for number in range(num + 1):
        if (number % 3 == 0) and number % 5 == 0:
            output = "FizzBuzz"
            print(f"{number} = {output}")
        elif number % 3 == 0:
            output = "Fizz"
            print(f"{number} = {output}")
        elif number % 5 == 0:
            output = "Buzz"
            print(f"{number} = {output}")
        else:
            print(number)


# fizz_buzz(50)


# Checking speed of Driver
# 0speed: The car Stop
# 1 - 70speed : OK
# >= 75speed - every 5km speed limit Demerit Points + 1
# 130 speed which is 12 demerit : License Suspended
def check_speed(speed):
    speed_limit = 70
    output = ""
    if speed == 0:
        output = "The car is stop"
    else:
        for num in range(speed):
            if speed <= speed_limit:
                output = "OK"
            elif speed > speed_limit:
                exceed_speed = speed - speed_limit
                per_speed_limit = exceed_speed / 5
                points = int(per_speed_limit)
                if points != 0 and points < 12:
                    output = f"Speed:{speed} Demerit Points:{points}"
                elif points >= 12:
                    output = f"License Suspended!!"
                else:
                    output = "OK"
    return output

# print(check_speed(130))


# Print ODD and EVEN beside numbers until limit reached
def show_numbers(limit):
    for num in range(limit + 1):
        if num % 2 == 0:
            print(f"{num} EVEN")
        else:
            print(f"{num} ODD")


# show_numbers(10)


# print Stars based on row input
def show_stars(rows):
    output = ""
    for rows in range(1, rows + 1):
        for cols in range(rows):
            output += "*"
        output += "\n"

    return output

# print(show_stars(5))


"""
Prime number is divisible by 2 natural numbers ONLY
- divide by 1 and itself
- greater than 1
- not more than 2 factors
- 2,3,5,7,11,13,17,19,23,29,31,37,41,43,47,53,59,61,67,71,73,79,83,89,97
factor : a number that divides another number exactly or evenly with no remainder
smallest prime is 2 
smallest composite number is 4

1 - not prime numbers because not > 1 not 2 factors
samples : 
                                no remainder
2 - 1,2             prime        2/1, 2/2
3 - 1,3             prime        3/1, 3/3
4 - 1,2,4           composite    4/1, 4/2, 4/4
5 - 1,5             prime        5/1, 5/5
6 - 1,2,3,6         composite    6/1, 6/2, 6/3, 6/6
7 - 1,7             prime        7/1, 7/7
8 - 1,2,4,8         composite    8/1, 8/2, 8/4, 8/8
9 - 1,3,9           composite    9/1, 9/3, 9/9
10 - 1,2,5          composite    10/1, 10/2, 10,5
11 - 1,11           prime        11/1, 11/11
12 - 1,2,3,4,6,12   composite    12/1, 12/2, 12/3, 12/4, 12/6, 12/12
13 - 1,13           prime        13/1, 13/13
14 - 1,2,7,14       composite
15 - 1,3,5,15       composite
18 - 1,2,3,6,9,18   composite
23 - 1,23           prime
29 - 1,29           prime

"""


# Get all Prime Numbers
def prime_numbers(num):
    count = 0
    prime = []
    for n in range(2, num+1):
        for in_n in range(1, num):
            if n % in_n == 0:
                count += 1
                if count == 2 and n == in_n:
                    prime.append(n)
        count = 0
    print(f'Prime Numbers : {prime}')


prime_numbers(100)
